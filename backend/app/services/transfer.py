"""Case export/import: a lossless, versioned ``.ecz`` (zip) bundle.

Export reads every case-scoped table verbatim and packs it with the original
attachment blobs and any managed source originals. Import reconstructs a brand
new case from such a bundle inside a single transaction, remapping every
integer id so the restored case is fully independent of the source database.

Design constraints honored here (see CLAUDE.md / improvementsv5.md):
- Emails are inserted column-for-column (NOT via repos.insert_email, which is
  ingest-tuned and would drop deleted_at/superseded_*/body_norm and emit a
  per-row audit event). We populate emails_fts explicitly so search works.
- The audit hash chain is GLOBAL; we never splice foreign events into the live
  audit_events table. Imported history is preserved inside the metadata of a
  single fresh ``case.imported`` event.
- Self-referential / cross-row foreign keys are nulled on first insert and
  fixed up in a second pass once the old->new id maps are complete.
"""

import io
import hashlib
import json
import sqlite3
import zipfile
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from .. import db
from ..config import ATTACHMENTS_DIR
from ..services.audit import record_audit_event
from ..services.text_norm import json_dumps, row_to_dict
from ..services import source_store
from ..repos import threads as thread_repo
from ..repos import duplicates as duplicate_repo

EXPORT_FORMAT = "emailchrono-export"
EXPORT_FORMAT_VERSION = 1


# --------------------------------------------------------------------------- #
# Shared helpers
# --------------------------------------------------------------------------- #

def max_schema_version(conn: sqlite3.Connection) -> Optional[str]:
    """Highest applied migration stem (e.g. "023_relative_paths").

    Version strings are zero-padded so lexical ordering matches numeric order.
    """
    row = conn.execute("SELECT MAX(version) AS v FROM schema_migrations").fetchone()
    return row["v"] if row else None


def _rows(conn: sqlite3.Connection, sql: str, params: Tuple = ()) -> List[Dict[str, Any]]:
    return [row_to_dict(r) for r in conn.execute(sql, params).fetchall()]


def _in_clause(values: List[Any]) -> str:
    return ",".join("?" for _ in values)


def _resolve_attachment_path(disk_path: str) -> Path:
    stored = Path(disk_path)
    return (stored if stored.is_absolute() else ATTACHMENTS_DIR / stored)


def safe_excel_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value)
    first = text.lstrip(" ")[:1]
    if first in {"=", "+", "-", "@", "\t", "\r", "\n"}:
        return f"'{text}"
    return text


# --------------------------------------------------------------------------- #
# Export
# --------------------------------------------------------------------------- #

class ExportError(Exception):
    pass


def export_case(case_id: int) -> bytes:
    """Build and return the full ``.ecz`` bundle for ``case_id`` as bytes."""
    with db.get_conn() as conn:
        case = conn.execute(
            "SELECT * FROM cases WHERE id = ?", (case_id,)
        ).fetchone()
        if not case:
            raise ExportError("Case not found")
        case_dict = row_to_dict(case)

        emails = _rows(conn, "SELECT * FROM emails WHERE case_id = ? ORDER BY id", (case_id,))
        email_ids = [e["id"] for e in emails]

        flags = attachments = snips = parts = []
        email_tags = []
        if email_ids:
            ph = _in_clause(email_ids)
            flags = _rows(conn, f"SELECT * FROM email_flags WHERE email_id IN ({ph}) ORDER BY id", tuple(email_ids))
            attachments = _rows(conn, f"SELECT * FROM attachments WHERE email_id IN ({ph}) ORDER BY id", tuple(email_ids))
            email_tags = _rows(conn, f"SELECT * FROM email_tags WHERE email_id IN ({ph})", tuple(email_ids))

        tags = _rows(conn, "SELECT * FROM tags WHERE case_id = ? ORDER BY id", (case_id,))
        snips = _rows(conn, "SELECT * FROM email_snips WHERE case_id = ? ORDER BY id", (case_id,))
        snip_ids = [s["id"] for s in snips]
        if snip_ids:
            ph = _in_clause(snip_ids)
            parts = _rows(conn, f"SELECT * FROM email_snip_parts WHERE snip_id IN ({ph}) ORDER BY id", tuple(snip_ids))
        text_mappings = _rows(conn, "SELECT * FROM email_text_mappings WHERE case_id = ? ORDER BY id", (case_id,))
        duplicates = _rows(conn, "SELECT * FROM email_duplicate_candidates WHERE case_id = ? ORDER BY id", (case_id,))
        jobs = _rows(conn, "SELECT * FROM ingest_jobs WHERE case_id = ? ORDER BY id", (case_id,))
        job_ids = [j["id"] for j in jobs]
        files = []
        if job_ids:
            ph = _in_clause(job_ids)
            files = _rows(conn, f"SELECT * FROM ingest_files WHERE job_id IN ({ph}) ORDER BY id", tuple(job_ids))
        case_folders = _rows(conn, "SELECT * FROM case_folders WHERE case_id = ? ORDER BY id", (case_id,))
        audit = _rows(conn, "SELECT * FROM audit_events WHERE case_id = ? ORDER BY id", (case_id,))

        manifest = {
            "format": EXPORT_FORMAT,
            "format_version": EXPORT_FORMAT_VERSION,
            "schema_version": max_schema_version(conn),
            "exported_at": db.utc_now(),
            "case_name": case_dict.get("name"),
            "counts": {
                "emails": len(emails),
                "attachments": len(attachments),
                "audit_events": len(audit),
            },
        }

    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("manifest.json", json_dumps(manifest))
        zf.writestr("case.json", json_dumps(case_dict))
        zf.writestr("emails.jsonl", "\n".join(json_dumps(e) for e in emails))
        zf.writestr("email_flags.json", json_dumps(flags))
        zf.writestr("tags.json", json_dumps({"tags": tags, "email_tags": email_tags}))
        zf.writestr("attachments.json", json_dumps(attachments))
        zf.writestr("snips.json", json_dumps({"snips": snips, "parts": parts}))
        zf.writestr("text_mappings.json", json_dumps(text_mappings))
        zf.writestr("duplicates.json", json_dumps(duplicates))
        zf.writestr("ingest.json", json_dumps({"jobs": jobs, "files": files}))
        zf.writestr("case_folders.json", json_dumps(case_folders))
        zf.writestr("audit.jsonl", "\n".join(json_dumps(a) for a in audit))

        # Attachment blobs, keyed by their stored (relative) disk_path.
        for att in attachments:
            disk_path = att.get("disk_path")
            if not disk_path:
                continue
            src = _resolve_attachment_path(disk_path)
            if src.exists() and src.is_file():
                arcname = f"attachments/{Path(disk_path).as_posix()}"
                zf.writestr(arcname, src.read_bytes())

        # Managed source originals, content-addressed.
        for shas in {e.get("source_blob_sha256") for e in emails if e.get("source_blob_sha256")}:
            blob = source_store.managed_path(shas)
            if blob.exists():
                zf.writestr(f"sources/{shas}", blob.read_bytes())

        xlsx = _build_chronology_xlsx(case_dict, emails, tags, email_tags)
        if xlsx is not None:
            zf.writestr("chronology.xlsx", xlsx)

    return buffer.getvalue()


def _build_chronology_xlsx(
    case_dict: Dict[str, Any],
    emails: List[Dict[str, Any]],
    tags: List[Dict[str, Any]],
    email_tags: List[Dict[str, Any]],
) -> Optional[bytes]:
    """Human-readable chronology sheet mirroring the frontend export columns.

    Best-effort: if openpyxl is unavailable the bundle simply omits the sheet
    (all machine-readable data is still present).
    """
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except Exception:
        return None

    tag_name = {t["id"]: t.get("name", "") for t in tags}
    tags_by_email: Dict[int, List[str]] = {}
    for et in email_tags:
        tags_by_email.setdefault(et["email_id"], []).append(tag_name.get(et["tag_id"], ""))

    headers = [
        "No", "Doc ID", "Date", "Thread", "From", "To", "Cc", "Subject", "Body",
        "Source", "Important", "Notes", "Confidence", "Flags", "Message ID", "Kind", "Tags",
    ]
    widths = [8, 12, 22, 12, 30, 34, 28, 42, 80, 48, 12, 42, 14, 30, 36, 14, 28]

    # Chronological view: visible rows only (skip soft-deleted / superseded).
    visible = [
        e for e in emails
        if not e.get("deleted_at") and not e.get("superseded_at")
    ]
    visible.sort(key=lambda e: (e.get("date_utc") or "", e.get("id")))

    thread_labels: Dict[str, int] = {}

    def _join_json(value: Any) -> str:
        try:
            parsed = json.loads(value) if isinstance(value, str) else value
        except (ValueError, TypeError):
            return ""
        if isinstance(parsed, list):
            return safe_excel_text("; ".join(str(x) for x in parsed))
        return safe_excel_text(parsed or "")

    wb = Workbook()
    ws = wb.active
    ws.title = "Chronology"
    ws.freeze_panes = "A2"
    ws.append(headers)
    for idx, email in enumerate(visible, start=1):
        key = email.get("thread_id") or str(email.get("id"))
        if key not in thread_labels:
            thread_labels[key] = len(thread_labels) + 1
        ws.append([
            idx,
            safe_excel_text(email.get("doc_id")),
            safe_excel_text(email.get("date_utc")),
            f"Thread {thread_labels[key]}",
            safe_excel_text(email.get("from_addr")),
            _join_json(email.get("to_json")),
            _join_json(email.get("cc_json")),
            safe_excel_text(email.get("subject")),
            safe_excel_text(email.get("body_text")),
            safe_excel_text(email.get("source_file_display")),
            "Yes" if email.get("important") else "No",
            safe_excel_text(email.get("notes")),
            email.get("parse_confidence"),
            _join_json(email.get("flags")) if email.get("flags") else "",
            safe_excel_text(email.get("message_id")),
            safe_excel_text(email.get("source_kind")),
            safe_excel_text("; ".join(tags_by_email.get(email["id"], []))),
        ])
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# --------------------------------------------------------------------------- #
# Import
# --------------------------------------------------------------------------- #

class ImportError_(Exception):
    """Raised on a malformed or incompatible bundle."""


def _validate_source_blobs(zf: zipfile.ZipFile) -> None:
    for name in zf.namelist():
        if name.startswith("sources\\"):
            raise ImportError_("Malformed source blob in bundle")
        if not name.startswith("sources/") or name.endswith("/"):
            continue
        sha = name.split("/", 1)[1]
        if (
            not sha
            or "/" in sha
            or "\\" in sha
            or ":" in sha
            or sha in {".", ".."}
            or not source_store.is_valid_sha256(sha)
        ):
            raise ImportError_("Malformed source blob in bundle")
        content = zf.read(name)
        if hashlib.sha256(content).hexdigest().lower() != sha.lower():
            raise ImportError_("Malformed source blob in bundle")


def _validate_email_source_refs(emails: List[Dict[str, Any]]) -> None:
    for email in emails:
        sha = email.get("source_blob_sha256")
        if sha and not source_store.is_valid_sha256(sha):
            raise ImportError_("Malformed source blob in bundle")


def _read_bundle(data: bytes) -> Dict[str, Any]:
    try:
        zf = zipfile.ZipFile(io.BytesIO(data))
    except zipfile.BadZipFile as exc:
        raise ImportError_("Not a valid .ecz archive") from exc

    def read_json(name: str, default: Any = None) -> Any:
        try:
            raw = zf.read(name).decode("utf-8")
        except KeyError:
            return default
        return json.loads(raw) if raw.strip() else default

    def read_jsonl(name: str) -> List[Any]:
        try:
            raw = zf.read(name).decode("utf-8")
        except KeyError:
            return []
        return [json.loads(line) for line in raw.splitlines() if line.strip()]

    manifest = read_json("manifest.json")
    if not manifest or manifest.get("format") != EXPORT_FORMAT:
        raise ImportError_("Missing or unrecognized manifest")

    tags_blob = read_json("tags.json", {}) or {}
    snips_blob = read_json("snips.json", {}) or {}
    ingest_blob = read_json("ingest.json", {}) or {}

    return {
        "zip": zf,
        "manifest": manifest,
        "case": read_json("case.json", {}) or {},
        "emails": read_jsonl("emails.jsonl"),
        "email_flags": read_json("email_flags.json", []) or [],
        "tags": tags_blob.get("tags", []),
        "email_tags": tags_blob.get("email_tags", []),
        "attachments": read_json("attachments.json", []) or [],
        "snips": snips_blob.get("snips", []),
        "snip_parts": snips_blob.get("parts", []),
        "text_mappings": read_json("text_mappings.json", []) or [],
        "duplicates": read_json("duplicates.json", []) or [],
        "ingest_jobs": ingest_blob.get("jobs", []),
        "ingest_files": ingest_blob.get("files", []),
        "case_folders": read_json("case_folders.json", []) or [],
        "audit": read_jsonl("audit.jsonl"),
    }


def preview_bundle(data: bytes) -> Dict[str, Any]:
    """Validate the manifest and report compatibility without writing."""
    bundle = _read_bundle(data)
    manifest = bundle["manifest"]
    with db.get_conn() as conn:
        current = max_schema_version(conn)
    export_schema = manifest.get("schema_version")
    incompatible = bool(export_schema and current and export_schema > current)
    refuse_reason = None
    if incompatible:
        refuse_reason = (
            f"Bundle schema {export_schema} is newer than this app ({current}); upgrade the app to import."
        )
    return {
        "format_version": manifest.get("format_version"),
        "schema_version": export_schema,
        "current_schema_version": current,
        "case_name": bundle["case"].get("name") or manifest.get("case_name"),
        "counts": {
            "emails": len(bundle["emails"]),
            "attachments": len(bundle["attachments"]),
            "audit_events": len(bundle["audit"]),
        },
        "compatible": not incompatible,
        "refuse_reason": refuse_reason,
    }


def _insert_row(conn: sqlite3.Connection, table: str, row: Dict[str, Any], drop: Tuple[str, ...] = ()) -> int:
    """Generic column-preserving insert. Returns the new rowid."""
    cols = [k for k in row.keys() if k not in drop]
    placeholders = ",".join("?" for _ in cols)
    sql = f"INSERT INTO {table} ({','.join(cols)}) VALUES ({placeholders})"
    cur = conn.execute(sql, tuple(row[c] for c in cols))
    return int(cur.lastrowid)


def import_bundle(data: bytes) -> Dict[str, Any]:
    """Reconstruct a new case from an ``.ecz`` bundle. Returns the new case row."""
    bundle = _read_bundle(data)
    manifest = bundle["manifest"]
    zf: zipfile.ZipFile = bundle["zip"]
    _validate_source_blobs(zf)
    _validate_email_source_refs(bundle["emails"])

    with db.get_conn() as conn:
        current = max_schema_version(conn)
        export_schema = manifest.get("schema_version")
        if export_schema and current and export_schema > current:
            raise ImportError_(
                f"Bundle schema {export_schema} is newer than this app ({current}); upgrade the app to import."
            )

        now = db.utc_now()

        # 1. New case row with a fresh public_id (avoid idx_cases_public_id clash).
        src_case = dict(bundle["case"])
        src_case.pop("id", None)
        src_case["public_id"] = _fresh_public_id(conn)
        src_case["created_at"] = src_case.get("created_at") or now
        src_case["updated_at"] = now
        case_id = _insert_row(conn, "cases", src_case)

        # 2. Emails, self/cross FKs nulled on first pass.
        SELF_FK = (
            "parent_email_id", "chain_source_id", "superseded_by_snip_id",
            "derived_from_attachment_id", "ingest_job_id",
        )
        email_map: Dict[int, int] = {}
        deferred_email_fks: Dict[int, Dict[str, Any]] = {}
        for e in bundle["emails"]:
            old_id = e["id"]
            row = dict(e)
            row.pop("id", None)
            row.pop("flags", None)  # GROUP_CONCAT artifact if present; not a column
            row["case_id"] = case_id
            saved = {fk: row.get(fk) for fk in SELF_FK}
            for fk in SELF_FK:
                row[fk] = None
            new_id = _insert_row(conn, "emails", row)
            email_map[old_id] = new_id
            deferred_email_fks[new_id] = saved
            # FTS row (search would silently break otherwise).
            conn.execute(
                "INSERT INTO emails_fts (rowid, subject, body_text) VALUES (?, ?, ?)",
                (new_id, row.get("subject"), row.get("body_text")),
            )

        # 3. Tags + remap email_tags.
        tag_map: Dict[int, int] = {}
        for t in bundle["tags"]:
            old_id = t["id"]
            row = dict(t)
            row.pop("id", None)
            row["case_id"] = case_id
            tag_map[old_id] = _insert_row(conn, "tags", row)

        # 4. Snips, then parts (depends on snip + email maps).
        snip_map: Dict[int, int] = {}
        for s in bundle["snips"]:
            old_id = s["id"]
            row = dict(s)
            row.pop("id", None)
            row["case_id"] = case_id
            row["source_email_id"] = email_map.get(row.get("source_email_id"))
            snip_map[old_id] = _insert_row(conn, "email_snips", row)

        # Second pass for email self/cross FKs now that all maps exist.
        for new_id, saved in deferred_email_fks.items():
            conn.execute(
                """
                UPDATE emails SET parent_email_id = ?, chain_source_id = ?,
                    superseded_by_snip_id = ?, derived_from_attachment_id = ?
                WHERE id = ?
                """,
                (
                    email_map.get(saved["parent_email_id"]),
                    email_map.get(saved["chain_source_id"]),
                    snip_map.get(saved["superseded_by_snip_id"]),
                    saved["derived_from_attachment_id"],  # remapped after attachments below
                    new_id,
                ),
            )

        # 5. email_flags.
        for f in bundle["email_flags"]:
            row = dict(f)
            row.pop("id", None)
            mapped = email_map.get(row.get("email_id"))
            if mapped is None:
                continue
            row["email_id"] = mapped
            _insert_row(conn, "email_flags", row)

        # 6. email_tags (composite, no id).
        for et in bundle["email_tags"]:
            row = dict(et)
            em = email_map.get(row.get("email_id"))
            tg = tag_map.get(row.get("tag_id"))
            if em is None or tg is None:
                continue
            row["email_id"] = em
            row["tag_id"] = tg
            _insert_row(conn, "email_tags", row)

        # 7. Attachments — copy blobs, recompute relative disk_path, remap.
        attachment_map: Dict[int, int] = {}
        for att in bundle["attachments"]:
            old_id = att["id"]
            old_email = att.get("email_id")
            new_email = email_map.get(old_email)
            if new_email is None:
                continue
            row = dict(att)
            row.pop("id", None)
            row["email_id"] = new_email
            old_disk = att.get("disk_path") or ""
            basename = Path(old_disk).name or f"{old_id}"
            new_rel = f"{new_email}/{basename}"
            arcname = f"attachments/{Path(old_disk).as_posix()}" if old_disk else None
            if arcname:
                try:
                    blob = zf.read(arcname)
                    target = ATTACHMENTS_DIR / new_rel
                    target.parent.mkdir(parents=True, exist_ok=True)
                    target.write_bytes(blob)
                except KeyError:
                    pass
            row["disk_path"] = new_rel
            attachment_map[old_id] = _insert_row(conn, "attachments", row)

        # Fix up emails.derived_from_attachment_id now attachment_map exists.
        for new_id, saved in deferred_email_fks.items():
            old_att = saved["derived_from_attachment_id"]
            if old_att is not None:
                conn.execute(
                    "UPDATE emails SET derived_from_attachment_id = ? WHERE id = ?",
                    (attachment_map.get(old_att), new_id),
                )

        # 8. Snip parts.
        for p in bundle["snip_parts"]:
            row = dict(p)
            row.pop("id", None)
            row["snip_id"] = snip_map.get(row.get("snip_id"))
            row["email_id"] = email_map.get(row.get("email_id"))
            if row["snip_id"] is None or row["email_id"] is None:
                continue
            _insert_row(conn, "email_snip_parts", row)

        # 9. Text mappings.
        for m in bundle["text_mappings"]:
            row = dict(m)
            row.pop("id", None)
            row["case_id"] = case_id
            row["source_email_id"] = email_map.get(row.get("source_email_id"))
            if "target_email_id" in row:
                row["target_email_id"] = email_map.get(row.get("target_email_id"))
            if row.get("snip_id") is not None:
                row["snip_id"] = snip_map.get(row.get("snip_id"))
            _insert_row(conn, "email_text_mappings", row)

        # 10. Duplicate candidates.
        for d in bundle["duplicates"]:
            row = dict(d)
            row.pop("id", None)
            row["case_id"] = case_id
            for col in ("email_a_id", "email_b_id", "canonical_email_id", "duplicate_email_id"):
                if row.get(col) is not None:
                    row[col] = email_map.get(row[col])
            if row.get("email_a_id") is None or row.get("email_b_id") is None:
                continue
            _insert_row(conn, "email_duplicate_candidates", row)

        # 11. Ingest jobs + files, then backfill emails.ingest_job_id.
        job_map: Dict[int, int] = {}
        for j in bundle["ingest_jobs"]:
            old_id = j["id"]
            row = dict(j)
            row.pop("id", None)
            row["case_id"] = case_id
            job_map[old_id] = _insert_row(conn, "ingest_jobs", row)
        for f in bundle["ingest_files"]:
            row = dict(f)
            row.pop("id", None)
            row["job_id"] = job_map.get(row.get("job_id"))
            if row.get("email_id") is not None:
                row["email_id"] = email_map.get(row["email_id"])
            if row["job_id"] is None:
                continue
            _insert_row(conn, "ingest_files", row)
        # emails.ingest_job_id second pass.
        for e in bundle["emails"]:
            old_job = e.get("ingest_job_id")
            if old_job is not None:
                conn.execute(
                    "UPDATE emails SET ingest_job_id = ? WHERE id = ?",
                    (job_map.get(old_job), email_map[e["id"]]),
                )

        # 12. Case folders.
        for cf in bundle["case_folders"]:
            row = dict(cf)
            row.pop("id", None)
            row["case_id"] = case_id
            _insert_row(conn, "case_folders", row)

        # 13. Copy managed source blobs.
        for name in zf.namelist():
            if name.startswith("sources/") and not name.endswith("/"):
                sha = name.split("/", 1)[1]
                source_store.store_source(zf.read(name), sha)

        # 14. Audit: do NOT splice foreign events into the global chain. Record
        # the imported history inside one fresh case.imported event's metadata.
        imported_audit = bundle["audit"]
        final_hash = imported_audit[-1].get("event_hash") if imported_audit else None
        record_audit_event(
            conn,
            case_id=case_id,
            action="case.imported",
            entity_type="case",
            entity_id=case_id,
            metadata={
                "format_version": manifest.get("format_version"),
                "schema_version": export_schema,
                "source_event_count": len(imported_audit),
                "source_final_event_hash": final_hash,
                "imported_audit_events": imported_audit,
            },
        )

        # 15. Recompute derived state (also backfills body_norm/minhash).
        thread_repo.recompute_case_threads(conn, case_id)

        conn.execute("UPDATE cases SET updated_at = ? WHERE id = ?", (now, case_id))
        new_case = row_to_dict(conn.execute("SELECT * FROM cases WHERE id = ?", (case_id,)).fetchone())

    # Duplicate recompute opens its own connection (post-commit), mirroring ingest.
    duplicate_repo.recompute_duplicate_candidates(case_id)
    return new_case


def _fresh_public_id(conn: sqlite3.Connection) -> str:
    import uuid

    for _ in range(50):
        candidate = uuid.uuid4().hex  # matches repos.cases.create_case convention
        exists = conn.execute(
            "SELECT 1 FROM cases WHERE public_id = ? LIMIT 1", (candidate,)
        ).fetchone()
        if not exists:
            return candidate
    return uuid.uuid4().hex
